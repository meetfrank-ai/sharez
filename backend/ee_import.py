"""
EasyEquities file import — parse transaction history XLSX,
store raw transactions, and derive holdings.
"""

import os
import re
import hashlib
import logging
from datetime import datetime, timezone
from io import BytesIO

from sqlalchemy.orm import Session

from models import User, Holding, UserTransaction, AccountType

logger = logging.getLogger(__name__)


def verify_ee_xlsx(file_bytes: bytes) -> dict:
    """Verify an XLSX file is genuinely from EasyEquities. Returns score and checks."""
    import openpyxl
    from datetime import datetime as dt

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
    except Exception:
        return {"score": 0, "max_score": 9, "checks": {}, "verified": False}

    ws = wb.active
    checks = {}

    # 1. Sheet name
    checks["sheet_name"] = ws.title == "Transaction History"

    # 2-4. Column headers
    checks["col_date"] = str(ws.cell(1, 1).value or "").strip() == "Date"
    checks["col_comment"] = str(ws.cell(1, 2).value or "").strip() == "Comment"
    checks["col_debit"] = str(ws.cell(1, 3).value or "").strip() == "Debit/Credit"

    # 5. Balance row
    row2_comment = str(ws.cell(2, 2).value or "")
    checks["balance_row"] = "Account Balance Carried Forward" in row2_comment

    # 6. Date is datetime object (not string)
    checks["date_is_datetime"] = isinstance(ws.cell(2, 1).value, dt)

    # 7-9. EE-specific fee patterns in first 100 rows
    has_broker_fee = False
    has_settlement = False
    has_vat = False
    for row in range(2, min(100, ws.max_row + 1)):
        comment = str(ws.cell(row, 2).value or "")
        if "Broker Commission" in comment:
            has_broker_fee = True
        if "Settlement and administration" in comment:
            has_settlement = True
        if "Value Added Tax" in comment:
            has_vat = True

    checks["has_broker_fees"] = has_broker_fee
    checks["has_settlement"] = has_settlement
    checks["has_vat"] = has_vat

    wb.close()

    score = sum(checks.values())
    max_score = len(checks)

    # Fee consistency check on first 3 buy transactions
    fee_consistent = _check_fee_consistency(file_bytes)
    checks["fee_consistency"] = fee_consistent
    if fee_consistent:
        score += 1
    max_score += 1

    verified = score >= 8  # 8/10 = verified, 5-7 = imported, <5 = unverified
    badge = "verified" if score >= 8 else "imported" if score >= 5 else "unverified"

    return {
        "score": score,
        "max_score": max_score,
        "checks": checks,
        "verified": verified,
        "badge": badge,
    }


def _check_fee_consistency(file_bytes: bytes) -> bool:
    """Check if broker fees are ~0.25% of trade amounts for first few buys."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
        ws = wb.active

        checked = 0
        consistent = 0

        for row in range(2, ws.max_row + 1):
            comment = str(ws.cell(row, 2).value or "")
            amount = ws.cell(row, 3).value or 0

            if comment.startswith("Bought ") and abs(amount) > 100:
                trade_amount = abs(amount)
                # Next row should be broker commission
                next_comment = str(ws.cell(row + 1, 2).value or "")
                next_amount = abs(ws.cell(row + 1, 3).value or 0)

                if "Broker Commission" in next_comment and next_amount > 0:
                    fee_pct = (next_amount / trade_amount) * 100
                    # EE charges 0.25% — allow 0.1% to 0.5% range
                    if 0.1 <= fee_pct <= 0.5:
                        consistent += 1
                    checked += 1

                if checked >= 3:
                    break

        wb.close()
        return checked > 0 and consistent == checked
    except Exception:
        return False


def parse_transaction_xlsx(file_bytes: bytes) -> dict:
    """Parse an EasyEquities Transaction History XLSX file."""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
    ws = wb.active

    transactions = []

    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row, 1).value
        comment = str(ws.cell(row, 2).value or '')
        amount = ws.cell(row, 3).value or 0

        match = re.match(r'(Bought|Sold) (.+?) ([\d,.]+) @ ([\d,.]+)', comment)
        if match:
            action = 'buy' if match.group(1) == 'Bought' else 'sell'
            stock_name = match.group(2).strip()
            quantity = float(match.group(3).replace(',', ''))
            price = float(match.group(4).replace(',', ''))

            # Create dedup hash from date + stock + action + quantity + price
            hash_input = f"{date_val}{stock_name}{action}{quantity}{price}"
            import_hash = hashlib.md5(hash_input.encode()).hexdigest()

            transactions.append({
                'date': str(date_val)[:19] if date_val else None,
                'action': action,
                'stock_name': stock_name,
                'quantity': quantity,
                'price': price,
                'amount': abs(amount),
                'import_hash': import_hash,
            })

    wb.close()

    # Aggregate for preview
    stocks = {}
    for t in transactions:
        name = t['stock_name']
        if name not in stocks:
            stocks[name] = {'stock_name': name, 'total_qty': 0, 'total_spent': 0, 'buy_count': 0, 'sell_count': 0}
        if t['action'] == 'buy':
            stocks[name]['total_qty'] += t['quantity']
            stocks[name]['total_spent'] += t['amount']
            stocks[name]['buy_count'] += 1
        else:
            stocks[name]['total_qty'] -= t['quantity']
            stocks[name]['sell_count'] += 1

    holdings = [
        {
            'stock_name': name,
            'quantity': round(data['total_qty'], 4),
            'total_invested': round(data['total_spent'], 2),
            'buy_count': data['buy_count'],
            'sell_count': data['sell_count'],
        }
        for name, data in sorted(stocks.items()) if data['total_qty'] > 0
    ]

    return {
        'transactions': transactions,
        'holdings': holdings,
        'total_stocks': len(holdings),
        'total_transactions': len(transactions),
    }


def import_transactions_to_db(
    db: Session,
    user: User,
    parsed_data: dict,
    account_type: str = "ZAR",
) -> dict:
    """Store raw transactions and rebuild holdings from them."""
    transactions = parsed_data.get('transactions', [])
    if not transactions:
        return {'message': 'No transactions found', 'count': 0, 'new_count': 0}

    new_count = 0
    for t in transactions:
        # Check dedup — skip if already imported
        existing = db.query(UserTransaction).filter(
            UserTransaction.user_id == user.id,
            UserTransaction.import_hash == t['import_hash'],
        ).first()
        if existing:
            continue

        # Generate contract code
        code = re.sub(r'[^A-Z0-9]', '', t['stock_name'].upper())[:10]
        contract_code = f"EE_{code}"

        # Parse date
        tx_date = None
        if t['date']:
            try:
                tx_date = datetime.strptime(t['date'][:19], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    tx_date = datetime.strptime(t['date'][:10], "%Y-%m-%d")
                except ValueError:
                    pass

        db.add(UserTransaction(
            user_id=user.id,
            action=t['action'],
            stock_name=t['stock_name'],
            contract_code=contract_code,
            account_type=account_type,
            quantity=t['quantity'],
            price=t['price'],
            amount=t['amount'],
            transaction_date=tx_date,
            import_hash=t['import_hash'],
        ))
        new_count += 1

    db.flush()

    # Rebuild holdings from all transactions for this account type
    _rebuild_holdings(db, user, account_type)

    db.commit()

    stocks = list(set(t['stock_name'] for t in transactions))
    return {
        'message': f'Imported {new_count} new transactions',
        'count': len(stocks),
        'new_count': new_count,
        'total_transactions': len(transactions),
        'stocks': sorted(stocks),
    }


def _rebuild_holdings(db: Session, user: User, account_type: str):
    """Derive holdings from all user transactions for an account type."""
    # Get all transactions for this account
    all_txs = (
        db.query(UserTransaction)
        .filter(UserTransaction.user_id == user.id, UserTransaction.account_type == account_type)
        .order_by(UserTransaction.transaction_date.asc())
        .all()
    )

    # Aggregate per stock + collect buy transactions for historical pricing
    stocks = {}
    buy_txs_by_stock = {}  # stock_name → [{date, amount}, ...]
    for tx in all_txs:
        name = tx.stock_name
        if name not in stocks:
            stocks[name] = {
                'stock_name': name,
                'contract_code': tx.contract_code,
                'shares': 0,
                'total_cost': 0,
            }
        if tx.action == 'buy':
            stocks[name]['shares'] += tx.quantity
            stocks[name]['total_cost'] += (tx.amount or 0)
            if name not in buy_txs_by_stock:
                buy_txs_by_stock[name] = []
            if tx.transaction_date and tx.amount:
                buy_txs_by_stock[name].append({
                    'date': tx.transaction_date,
                    'amount': tx.amount,
                })
        else:
            stocks[name]['shares'] -= tx.quantity

    # Map account type
    try:
        acc_type = AccountType(account_type)
    except ValueError:
        acc_type = AccountType.ZAR

    # Clear existing holdings for this account type and rebuild
    db.query(Holding).filter(
        Holding.user_id == user.id,
        Holding.account_type == acc_type,
    ).delete()

    now = datetime.now(timezone.utc)

    for name, data in stocks.items():
        if data['shares'] <= 0:
            continue  # Fully sold, don't create holding

        avg_price = data['total_cost'] / data['shares'] if data['shares'] > 0 else 0

        db.add(Holding(
            user_id=user.id,
            account_type=acc_type,
            stock_name=name,
            contract_code=data['contract_code'],
            purchase_value=data['total_cost'],
            current_value=data['total_cost'],  # Will be updated by refresh_holdings_prices
            current_price=avg_price,
            shares=data['shares'],
            last_synced_at=now,
        ))

    # Auto-map any unmapped holdings to EODHD tickers
    try:
        _auto_map_instruments(db, stocks.keys())
    except Exception as e:
        logger.warning(f"Auto-map instruments failed: {e}")

    # Fetch historical prices for buy transactions → compute external_avg_buy_price
    try:
        _compute_external_buy_prices(db, user, buy_txs_by_stock)
    except Exception as e:
        logger.warning(f"Historical price fetch failed: {e}")

    # Try to update with live prices
    try:
        _refresh_holdings_prices(db, user)
    except Exception as e:
        logger.warning(f"Price refresh failed for user {user.id}: {e}")


def _compute_external_buy_prices(db: Session, user: User, buy_txs_by_stock: dict):
    """Fetch historical prices for each buy transaction and compute weighted avg external buy price."""
    import time
    from models import InstrumentMap
    from price_resolver import resolve_historical_price

    holdings = db.query(Holding).filter(Holding.user_id == user.id).all()
    holdings_by_name = {h.stock_name: h for h in holdings}

    fetched = 0
    for stock_name, txs in buy_txs_by_stock.items():
        holding = holdings_by_name.get(stock_name)
        if not holding or not txs:
            continue

        # Look up instrument mapping for EODHD symbol
        mapping = db.query(InstrumentMap).filter(InstrumentMap.ee_name == stock_name).first()
        if not mapping or not mapping.eodhd_symbol:
            continue

        # Fetch historical price for each buy transaction date
        total_weighted_price = 0
        total_amount = 0
        resolved_count = 0

        for tx in txs:
            result = resolve_historical_price(
                db, mapping.eodhd_symbol, tx['date'],
                yf_symbol=mapping.yfinance_symbol,
            )
            if result["price"]:
                total_weighted_price += tx['amount'] * result["price"]
                total_amount += tx['amount']
                resolved_count += 1

            fetched += 1
            if fetched % 5 == 0:
                time.sleep(0.2)  # Rate limit

        if total_amount > 0 and resolved_count > 0:
            # Weighted average: sum(amount_i × price_i) / sum(amount_i)
            holding.external_avg_buy_price = round(total_weighted_price / total_amount, 4)
            logger.info(f"External avg buy for {stock_name}: R{holding.external_avg_buy_price:.4f} ({resolved_count}/{len(txs)} trades resolved)")

    db.flush()


def _refresh_holdings_prices(db: Session, user: User):
    """Refresh current prices and calculate P&L using external historical prices where available."""
    from price_resolver import resolve_price

    holdings = db.query(Holding).filter(Holding.user_id == user.id).all()
    now = datetime.now(timezone.utc)

    for h in holdings:
        # Use external_avg_buy_price for sanity check if available, else fall back to EE avg
        avg_price = h.external_avg_buy_price
        if not avg_price:
            avg_price = h.purchase_value / h.shares if (h.shares and h.shares > 0 and h.purchase_value) else None

        result = resolve_price(db, h.stock_name, avg_buy_price=avg_price)

        if result["price"] and result["source"] != "none":
            h.current_price = result["price"]
            h.price_source = result["source"]
            h.last_synced_at = now

            if h.external_avg_buy_price and h.external_avg_buy_price > 0:
                # Use external prices for P&L — same scale, correct calculation
                pnl_pct = (result["price"] - h.external_avg_buy_price) / h.external_avg_buy_price
                h.current_value = round(h.purchase_value * (1 + pnl_pct), 2)
                logger.info(f"Updated {h.stock_name}: ext_buy=R{h.external_avg_buy_price:.2f} curr=R{result['price']:.2f} P&L={pnl_pct*100:+.1f}% [{result['source']}]")
            elif not result.get("unit_mismatch"):
                # No external buy price, not a unit trust — use old method
                if h.shares and h.shares > 0:
                    h.current_value = round(h.shares * result["price"], 2)
                logger.info(f"Updated {h.stock_name}: R{result['price']:.2f} (no ext buy, using shares×price) [{result['source']}]")
            else:
                # Unit trust with unit mismatch — keep purchase_value as current_value
                logger.info(f"NAV ref for {h.stock_name}: R{result['price']:.2f} (unit mismatch) [{result['source']}]")
        else:
            logger.info(f"No price available for {h.stock_name}")

    db.commit()


def _auto_map_instruments(db: Session, stock_names):
    """Auto-create instrument_map entries for unmapped stocks using JSE symbol list."""
    from models import InstrumentMap
    from ticker_resolver import resolve_ticker, get_jse_symbols

    jse_symbols = get_jse_symbols()
    jse_by_code = {s.get("Code", "").upper(): s for s in jse_symbols}

    mapped = 0
    for name in stock_names:
        # Skip if already mapped
        existing = db.query(InstrumentMap).filter(InstrumentMap.ee_name == name).first()
        if existing:
            continue

        ticker = resolve_ticker(name)
        if not ticker:
            logger.info(f"Auto-map: no ticker found for '{name}'")
            continue

        # ticker is like "PRX.JSE" — extract code
        code = ticker.replace(".JSE", "")
        jse_info = jse_by_code.get(code, {})

        # Determine instrument type from EODHD type field
        eodhd_type = jse_info.get("Type", "").lower()
        if "etf" in eodhd_type or "fund" in eodhd_type:
            inst_type = "etf"
        elif "ametf" in name.lower():
            inst_type = "ametf"
        elif any(kw in name.lower() for kw in ("fund", "trust", "bci", "sci")):
            inst_type = "unit_trust"
        else:
            inst_type = "stock"

        yf_symbol = f"{code}.JO"  # yfinance uses .JO for JSE

        db.add(InstrumentMap(
            ee_name=name,
            ticker=code,
            market="JSE",
            instrument_type=inst_type,
            eodhd_symbol=ticker,
            yfinance_symbol=yf_symbol,
            sector=jse_info.get("Sector", None),
            is_verified=False,  # auto-mapped, not manually verified
        ))
        mapped += 1
        logger.info(f"Auto-mapped: '{name}' → {ticker} ({inst_type})")

    if mapped:
        db.flush()
        logger.info(f"Auto-mapped {mapped} new instruments")


def refresh_user_prices(db: Session, user: User):
    """Public function to refresh holdings prices for a user."""
    _refresh_holdings_prices(db, user)
